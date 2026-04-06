"""
Database cleanup service.

Fixes problems introduced by bulk ОСВ import:
  1. Merges duplicate numbered nomenclatures (Название 1, Название 2 → Название)
  2. Regenerates IMP-XXXXXX SKUs to proper Cyrillic-prefix format
  3. Fixes wrong accounting_type (счёт 01 items must be 'individual')
  4. Imports real inventory data from data/Инвенторизация/ files
"""
import os
import re
import random
import logging
from datetime import datetime

logger = logging.getLogger('services.db_cleanup')

INV_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data', 'Инвенторизация')


# ---------------------------------------------------------------------------
# SKU generator (shared with excel_import)
# ---------------------------------------------------------------------------

def generate_sku_from_name(name: str, existing_skus: set) -> str:
    """
    Generates SKU like existing examples:
      "Сапоги резиновые утепленные мужский" → "САП-РЕЗ-УТЕ-724569"
      "Берцы демисезонные"                  → "БЕР-ДЕМ-742890"
      "Комплект термобелья"                 → "КОМ-ТЕР-744782"
      "Повербанк"                           → "ПОВ-817700"

    Algorithm: take first 3 chars of each word with >=3 chars, up to 3 words,
    join with '-', append '-' + 6-digit random number.
    """
    clean = re.sub(r'[^\w\s]', ' ', name)
    words = clean.split()

    parts = []
    for word in words:
        if len(word) >= 3:
            parts.append(word[:3].upper())
        if len(parts) >= 3:
            break

    if not parts:
        base = name[:3].upper() if len(name) >= 3 else (name + 'XXX')[:3].upper()
        parts = [base]

    prefix = '-'.join(parts)
    for _ in range(200):
        suffix = str(random.randint(100000, 999999))
        sku = f"{prefix}-{suffix}"
        if sku not in existing_skus:
            existing_skus.add(sku)
            return sku
    # fallback (extremely rare)
    return f"{prefix}-{random.randint(100000, 999999)}"


# ---------------------------------------------------------------------------
# Main cleanup class
# ---------------------------------------------------------------------------

class DatabaseCleanup:
    def __init__(self, db_path='data/assets.db'):
        import sqlite3
        self.conn = sqlite3.connect(db_path)
        self.conn.row_factory = sqlite3.Row
        self.conn.execute("PRAGMA foreign_keys = OFF")
        self._existing_skus: set = set()
        self._nom_cache: dict = {}   # normalised_name_lower -> canonical_id (built during merge)
        self.stats = {
            'nomenclatures_merged': 0,
            'nomenclatures_renamed': 0,
            'skus_regenerated': 0,
            'accounting_types_fixed': 0,
            'instances_moved': 0,
            'instances_from_inventory': 0,
            'nomenclatures_from_inventory': 0,
            'errors': [],
        }

    # ------------------------------------------------------------------
    def run(self):
        """Run all cleanup steps in order."""
        logger.info("Starting DB cleanup …")
        self._load_existing_skus()
        self.merge_numbered_duplicates()
        self.fix_accounting_types()
        self.regenerate_imp_skus()
        self.import_inventory_files()
        self.conn.execute("PRAGMA foreign_keys = ON")
        self.conn.close()
        logger.info("DB cleanup finished: %s", self.stats)
        return self.stats

    # ------------------------------------------------------------------
    def _load_existing_skus(self):
        c = self.conn.cursor()
        c.execute("SELECT sku FROM nomenclatures WHERE sku IS NOT NULL")
        self._existing_skus = {r['sku'] for r in c.fetchall()}
        c.close()
        logger.info("Loaded %d existing SKUs", len(self._existing_skus))

    # ------------------------------------------------------------------
    def merge_numbered_duplicates(self):
        """
        Normalises names like "Название 1", "Название 2" → "Название".
        Merges instances/stocks onto the canonical record, soft-deletes duplicates.
        """
        logger.info("merge_numbered_duplicates …")
        c = self.conn.cursor()
        c.execute(
            "SELECT id, name, sku, accounting_type FROM nomenclatures "
            "WHERE (is_deleted = 0 OR is_deleted IS NULL)"
        )
        rows = c.fetchall()

        # Group by normalised name
        from collections import defaultdict
        groups: dict = defaultdict(list)
        for row in rows:
            norm = re.sub(r'\s*\.?\s*\d+\s*$', '', row['name']).strip()
            groups[norm].append(row)

        for norm, members in groups.items():
            if len(members) == 1:
                # Single entry — just add to cache
                self._nom_cache[norm.lower()] = members[0]['id']
                continue

            # Find canonical: exact name match first, else lowest id
            exact = [m for m in members if m['name'] == norm]
            if exact:
                canonical = exact[0]
            else:
                canonical = min(members, key=lambda m: m['id'])

            duplicates = [m for m in members if m['id'] != canonical['id']]

            # If no exact match exists, rename canonical to base name
            if not exact:
                c.execute(
                    "UPDATE nomenclatures SET name=? WHERE id=?",
                    (norm, canonical['id'])
                )
                self.stats['nomenclatures_renamed'] += 1
                logger.debug("Renamed id=%d to '%s'", canonical['id'], norm)

            for dup in duplicates:
                dup_id = dup['id']
                can_id = canonical['id']

                # Move instances
                c.execute(
                    "UPDATE instances SET nomenclature_id=? WHERE nomenclature_id=?",
                    (can_id, dup_id)
                )
                moved = c.rowcount
                self.stats['instances_moved'] += moved

                # Move stocks (handle unique constraint)
                c.execute(
                    "SELECT id, warehouse_id, storage_bin_id, batch_id, quantity "
                    "FROM stocks WHERE nomenclature_id=?", (dup_id,)
                )
                dup_stocks = c.fetchall()
                for ds in dup_stocks:
                    c.execute(
                        "SELECT id, quantity FROM stocks "
                        "WHERE nomenclature_id=? AND warehouse_id=? "
                        "  AND (storage_bin_id IS ? OR storage_bin_id=?) "
                        "  AND (batch_id IS ? OR batch_id=?)",
                        (can_id,
                         ds['warehouse_id'],
                         ds['storage_bin_id'], ds['storage_bin_id'],
                         ds['batch_id'], ds['batch_id'])
                    )
                    existing = c.fetchone()
                    if existing:
                        c.execute(
                            "UPDATE stocks SET quantity=quantity+? WHERE id=?",
                            (ds['quantity'], existing['id'])
                        )
                        c.execute("DELETE FROM stocks WHERE id=?", (ds['id'],))
                    else:
                        c.execute(
                            "UPDATE stocks SET nomenclature_id=? WHERE id=?",
                            (can_id, ds['id'])
                        )

                # Soft-delete duplicate
                c.execute(
                    "UPDATE nomenclatures SET is_deleted=1 WHERE id=?", (dup_id,)
                )
                self.stats['nomenclatures_merged'] += 1

            self._nom_cache[norm.lower()] = canonical['id']

        self.conn.commit()
        c.close()
        logger.info(
            "merge_numbered_duplicates: merged=%d renamed=%d moved_instances=%d",
            self.stats['nomenclatures_merged'],
            self.stats['nomenclatures_renamed'],
            self.stats['instances_moved'],
        )

    # ------------------------------------------------------------------
    def fix_accounting_types(self):
        """
        1. quantitative items that already have instances → individual
        2. PPE keyword items → individual + has_serial_numbers=1
        """
        logger.info("fix_accounting_types …")
        c = self.conn.cursor()

        # Fix 1: quantitative with instances → individual
        c.execute("""
            SELECT DISTINCT n.id, n.name, n.accounting_type
            FROM nomenclatures n
            JOIN instances i ON i.nomenclature_id = n.id
            WHERE n.accounting_type = 'quantitative'
              AND (n.is_deleted = 0 OR n.is_deleted IS NULL)
        """)
        for row in c.fetchall():
            c.execute(
                "UPDATE nomenclatures SET accounting_type='individual' WHERE id=?",
                (row['id'],)
            )
            self.stats['accounting_types_fixed'] += 1

        # Fix 2: PPE keywords → individual
        ppe_keywords = [
            'сапог', 'берц', 'перчат', 'каска', 'жилет', 'костюм',
            'защит', 'маска', 'наушник', 'ботин', 'кроссовк',
        ]
        for kw in ppe_keywords:
            c.execute("""
                UPDATE nomenclatures
                   SET accounting_type='individual', has_serial_numbers=1
                 WHERE LOWER(name) LIKE ?
                   AND accounting_type != 'individual'
                   AND (is_deleted = 0 OR is_deleted IS NULL)
            """, (f'%{kw}%',))
            self.stats['accounting_types_fixed'] += c.rowcount

        self.conn.commit()
        c.close()
        logger.info("fix_accounting_types: fixed=%d", self.stats['accounting_types_fixed'])

    # ------------------------------------------------------------------
    def regenerate_imp_skus(self):
        """Replace IMP-XXXXXX SKUs with proper Cyrillic-prefix SKUs."""
        logger.info("regenerate_imp_skus …")
        c = self.conn.cursor()
        c.execute(
            "SELECT id, name FROM nomenclatures "
            "WHERE sku LIKE 'IMP-%' AND (is_deleted=0 OR is_deleted IS NULL)"
        )
        rows = c.fetchall()
        logger.info("  Found %d IMP SKUs to regenerate", len(rows))
        for row in rows:
            new_sku = generate_sku_from_name(row['name'], self._existing_skus)
            c.execute("UPDATE nomenclatures SET sku=? WHERE id=?", (new_sku, row['id']))
            self.stats['skus_regenerated'] += 1
        self.conn.commit()
        c.close()
        logger.info("regenerate_imp_skus: regenerated=%d", self.stats['skus_regenerated'])

    # ------------------------------------------------------------------
    def import_inventory_files(self):
        """
        Import real inventory data from data/Инвенторизация/.
        Priority: Экземпляры.xlsx first (most recent), then 30.11.25_v2.xlsx,
        then 30.11.25 инвентаризация .xlsx.
        """
        logger.info("import_inventory_files …")

        # Pre-load existing inventory numbers to avoid duplicates
        c = self.conn.cursor()
        c.execute("SELECT inventory_number FROM instances WHERE inventory_number IS NOT NULL")
        self._inv_set: set = {r['inventory_number'] for r in c.fetchall()}
        c.close()

        # Load default warehouse
        c = self.conn.cursor()
        c.execute("SELECT id FROM warehouses WHERE is_active=1 ORDER BY id LIMIT 1")
        row = c.fetchone()
        self._default_warehouse_id = row['id'] if row else 1
        c.close()

        # Load default category
        c = self.conn.cursor()
        c.execute("SELECT id FROM categories WHERE parent_id IS NULL ORDER BY id LIMIT 1")
        row = c.fetchone()
        self._default_category_id = row['id'] if row else 1
        c.close()

        # Build name→id cache from current (post-merge) nomenclatures
        c = self.conn.cursor()
        c.execute(
            "SELECT id, name FROM nomenclatures WHERE is_deleted=0 OR is_deleted IS NULL"
        )
        self._inv_nom_cache: dict = {}
        for row in c.fetchall():
            self._inv_nom_cache[row['name'].strip().lower()] = row['id']
        c.close()

        files = [
            ('Экземпляры.xlsx', 'instances'),
            ('30.11.25_v2.xlsx', 'inventory'),
            ('30.11.25 инвентаризация .xlsx', 'inventory'),
        ]

        for fname, fmt in files:
            path = os.path.join(INV_DIR, fname)
            if not os.path.exists(path):
                logger.warning("Inventory file not found: %s", fname)
                continue
            try:
                if fmt == 'instances':
                    self._import_instances_file(path)
                else:
                    self._import_inventory_file(path)
                logger.info("  Imported: %s", fname)
            except Exception as e:
                import traceback
                msg = f"{fname}: {e}\n{traceback.format_exc()}"
                logger.error(msg)
                self.stats['errors'].append(msg)

        logger.info(
            "import_inventory_files: instances=%d nomenclatures=%d",
            self.stats['instances_from_inventory'],
            self.stats['nomenclatures_from_inventory'],
        )

    # ------------------------------------------------------------------
    # Экземпляры.xlsx importer
    # ------------------------------------------------------------------

    def _import_instances_file(self, path: str):
        """
        Sheet 'Шаблон':
          0  Старый инв. номер
          1  Инв. номер          ← dedup key
          2  Номенклатура/Комплект ← name
          3  Комплектующие       (skip)
          4  Серийный номер
          5  Статус
          6  Местоположение
          7  Сотрудник
          8  Дата покупки
          9  Цена
          10 Количество
        """
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet_name = 'Шаблон' if 'Шаблон' in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        for row in rows:
            try:
                if not row or not row[2]:
                    continue
                old_inv = str(row[0]).strip() if row[0] else None
                inv_num = str(row[1]).strip() if row[1] else None
                name = str(row[2]).strip()
                serial = str(row[4]).strip() if row[4] else None
                status = str(row[5]).strip() if row[5] else 'in_stock'
                location = str(row[6]).strip() if row[6] else None
                employee = str(row[7]).strip() if row[7] else None
                purchase_date = row[8]
                price = row[9]

                if inv_num and inv_num in self._inv_set:
                    continue
                if not inv_num:
                    continue

                nom_id = self._get_or_create_nomenclature(name)
                self._create_inv_instance(
                    nom_id=nom_id,
                    inv_num=inv_num,
                    old_inv=old_inv,
                    serial=serial,
                    status=status,
                    location=location,
                    employee=employee,
                    purchase_date=purchase_date,
                    price=price,
                )
            except Exception as e:
                self.stats['errors'].append(f"Экземпляры row: {e}")

    # ------------------------------------------------------------------
    # Inventory xlsx importer (30.11.25 format)
    # ------------------------------------------------------------------

    def _import_inventory_file(self, path: str):
        """
        Single sheet, columns:
          0  №
          1  Наименование         ← main item name
          2  Комплектация         ← sub-item name (when col0/1 == 'в том числе')
          3  кол./фактическое
          4  инвентарный номер    ← dedup key
          5  год
          6  Стоимость
          7  Поставщик
          8  Факт
          9  Примечание
        """
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        current_main_name = None

        for row in rows:
            try:
                if not row or all(c is None for c in row[:5]):
                    continue

                col0 = str(row[0]).strip() if row[0] is not None else ''
                col1 = str(row[1]).strip() if row[1] is not None else ''

                is_sub = (col0.lower() == 'в том числе' or col1.lower() == 'в том числе')

                if is_sub:
                    # sub-item: name in col2, inv in col4
                    name_raw = str(row[2]).strip() if row[2] else None
                    if not name_raw:
                        continue
                    name = name_raw
                else:
                    # Main item row: number in col0, name in col1
                    if not col1 or col1.lower() in ('наименование', 'в том числе', ''):
                        continue
                    # Skip header-like rows
                    try:
                        float(col0)
                    except (ValueError, TypeError):
                        continue
                    name = col1
                    current_main_name = name

                inv_num = str(row[4]).strip() if row[4] else None
                price = row[6]
                supplier_name = str(row[7]).strip() if row[7] else None

                if not inv_num:
                    continue
                if inv_num in self._inv_set:
                    continue

                nom_id = self._get_or_create_nomenclature(name)
                supplier_id = self._get_or_create_supplier(supplier_name) if supplier_name else None

                self._create_inv_instance(
                    nom_id=nom_id,
                    inv_num=inv_num,
                    old_inv=None,
                    serial=None,
                    status='in_stock',
                    location=None,
                    employee=None,
                    purchase_date=None,
                    price=price,
                    supplier_id=supplier_id,
                )
            except Exception as e:
                self.stats['errors'].append(f"{os.path.basename(path)} row: {e}")

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _get_or_create_nomenclature(self, name: str) -> int:
        key = name.strip().lower()
        if key in self._inv_nom_cache:
            return self._inv_nom_cache[key]
        # Also check _nom_cache (from merge step)
        if key in self._nom_cache:
            self._inv_nom_cache[key] = self._nom_cache[key]
            return self._nom_cache[key]

        sku = generate_sku_from_name(name, self._existing_skus)
        c = self.conn.cursor()
        c.execute("""
            INSERT INTO nomenclatures
                (sku, name, category_id, accounting_type, unit,
                 has_serial_numbers, is_active, is_deleted, created_at)
            VALUES (?, ?, ?, 'individual', 'шт', 1, 1, 0, CURRENT_TIMESTAMP)
        """, (sku, name.strip(), self._default_category_id))
        nom_id = c.lastrowid
        self.conn.commit()
        c.close()

        self._inv_nom_cache[key] = nom_id
        self.stats['nomenclatures_from_inventory'] += 1
        return nom_id

    def _get_or_create_supplier(self, name: str) -> int:
        """Find or create supplier by name. Returns supplier id."""
        if not hasattr(self, '_supplier_cache'):
            c = self.conn.cursor()
            c.execute("SELECT id, name FROM suppliers")
            self._supplier_cache = {r['name'].strip().lower(): r['id'] for r in c.fetchall()}
            c.close()
        key = name.strip().lower()
        if key in self._supplier_cache:
            return self._supplier_cache[key]

        # Generate a unique code from the name
        raw_code = re.sub(r'[^A-Za-zА-Яа-я0-9]', '', name).upper()[:20] or 'SUP'
        code = raw_code
        c = self.conn.cursor()
        c.execute("SELECT COUNT(*) FROM suppliers WHERE code LIKE ?", (raw_code + '%',))
        cnt = c.fetchone()[0]
        if cnt:
            code = f"{raw_code[:16]}{cnt:04d}"
        try:
            c.execute(
                "INSERT INTO suppliers (code, name, created_at) VALUES (?, ?, CURRENT_TIMESTAMP)",
                (code, name.strip())
            )
            sid = c.lastrowid
            self.conn.commit()
        except Exception:
            self.conn.rollback()
            # If code collision, try with a suffix
            import random as _rnd
            code = f"{raw_code[:14]}{_rnd.randint(1000,9999)}"
            c.execute(
                "INSERT OR IGNORE INTO suppliers (code, name, created_at) VALUES (?, ?, CURRENT_TIMESTAMP)",
                (code, name.strip())
            )
            sid = c.lastrowid
            self.conn.commit()
        c.close()
        self._supplier_cache[key] = sid
        return sid

    STATUS_MAP = {
        'in_stock': 'in_stock', 'в наличии': 'in_stock',
        'active': 'in_stock',
        'written_off': 'written_off', 'списан': 'written_off',
        'repair': 'repair', 'ремонт': 'repair',
        'in_use': 'in_use', 'в использовании': 'in_use', 'выдан': 'in_use',
    }

    def _create_inv_instance(
        self,
        nom_id: int,
        inv_num: str,
        old_inv: str,
        serial: str,
        status: str,
        location: str,
        employee: str,
        purchase_date,
        price,
        supplier_id: int = None,
    ):
        if not nom_id or not inv_num:
            return
        if inv_num in self._inv_set:
            return

        mapped_status = self.STATUS_MAP.get(
            status.lower() if status else '', 'in_stock'
        )

        # Parse price
        price_val = None
        if price:
            try:
                price_val = float(str(price).replace(',', '.').replace(' ', '').replace('\xa0', ''))
            except (ValueError, TypeError):
                price_val = None

        # Parse date
        pd_str = None
        if purchase_date:
            if isinstance(purchase_date, datetime):
                pd_str = purchase_date.strftime('%Y-%m-%d')
            else:
                try:
                    pd_str = str(purchase_date)[:10]
                except Exception:
                    pd_str = None

        c = self.conn.cursor()
        try:
            c.execute("""
                INSERT INTO instances
                    (nomenclature_id, inventory_number, old_inventory_number,
                     serial_number, status, condition, supplier_id,
                     purchase_date, purchase_price, created_at)
                VALUES (?, ?, ?, ?, ?, 'good', ?, ?, ?, CURRENT_TIMESTAMP)
            """, (nom_id, inv_num, old_inv, serial, mapped_status,
                  supplier_id, pd_str, price_val))
            self.conn.commit()
            self._inv_set.add(inv_num)
            self.stats['instances_from_inventory'] += 1
        except Exception as e:
            self.conn.rollback()
            self.stats['errors'].append(f"create_inv_instance {inv_num}: {e}")
        finally:
            c.close()
