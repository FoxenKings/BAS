"""
Excel import service: reads data files from data/ folder,
creates nomenclature entries and instances, skips duplicates.
"""
import os
import re
import logging
from datetime import datetime
import openpyxl
from services.db_cleanup import generate_sku_from_name

logger = logging.getLogger('services.excel_import')

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data')

STATUS_MAP = {
    'active': 'in_stock',
    'in_stock': 'in_stock',
    'В наличии': 'in_stock',
    'written_off': 'written_off',
    'Списан': 'written_off',
    'repair': 'repair',
    'Ремонт': 'repair',
    'in_use': 'in_use',
    'В использовании': 'in_use',
    'Выдан': 'in_use',
}


class ExcelImportService:
    def __init__(self, db, user_id=1):
        self.db = db
        self.user_id = user_id
        self.stats = {
            'nomenclatures_created': 0,
            'nomenclatures_skipped': 0,
            'instances_created': 0,
            'instances_skipped': 0,
            'errors': [],
            'files_processed': []
        }
        self._nom_cache = {}   # name_lower -> nomenclature_id
        self._inv_cache = set()    # existing inventory_numbers
        self._serial_cache = set()  # existing serial_numbers
        self._existing_skus = set()  # existing SKUs (for generate_sku_from_name)

    def run(self):
        """Run full import from all Excel files."""
        try:
            self._load_existing_data()
            self._default_category_id = self._get_default_category_id()
            self._category_id_01 = self._get_category_id_by_code('ASSETS', fallback=1)
            self._category_id_10 = self._get_category_id_by_code('INVENTORY', fallback=2)
            self._default_warehouse_id = self._get_default_warehouse_id()
            self._stock_cache = self._load_stock_cache()

            files = [
                ('inventory.xlsx', 'standard'),
                ('equipment.xlsx', 'standard'),
                ('assets.xlsx', 'standard'),
                ('consumable.xlsx', 'standard'),
                ('tool.xlsx', 'tool'),
                ('Экземпляры.xlsx', 'instances'),
                # ОСВ счёт 01 — основные средства (newest first for dedup priority)
                ('Оборотно_сальдовая_ведомость_по_счету_01_за_1_квартал_2026_г_АНО.xlsx', 'osv01'),
                ('Оборотно_сальдовая_ведомость_по_счету_01_за_2025_г_АНО_«НПЦ_БАС.xlsx', 'osv01'),
                ('Оборотно_сальдовая_ведомость_по_счету_01_за_2024_г_АНО_«НПЦ_БАС.xlsx', 'osv01'),
                # ОСВ счёт 10 — материалы (newest first)
                ('Оборотно_сальдовая_ведомость_по_счету_10_за_1_квартал_2026_г_АНО.xlsx', 'osv10'),
                ('Оборотно_сальдовая_ведомость_по_счету_10_за_2025_г_АНО_«НПЦ_БАС.xlsx', 'osv10'),
            ]

            for filename, fmt in files:
                path = os.path.join(DATA_DIR, filename)
                if not os.path.exists(path):
                    logger.warning(f"File not found: {filename}")
                    continue
                try:
                    if fmt in ('osv01', 'osv10'):
                        self._import_osv_file(path, fmt)
                    else:
                        self._import_file(path, fmt)
                    self.stats['files_processed'].append(filename)
                except Exception as e:
                    logger.error(f"Error importing {filename}: {e}")
                    self.stats['errors'].append(f"{filename}: {str(e)}")
        finally:
            pass

        return self.stats

    def _get_default_category_id(self) -> int:
        """Return id of the first root category as fallback."""
        row = self.db.execute_query(
            "SELECT id FROM categories WHERE parent_id IS NULL ORDER BY id LIMIT 1",
            fetch_all=False
        )
        return row['id'] if row else 1

    def _load_existing_data(self):
        """Cache existing inventory numbers and serial numbers to detect duplicates."""
        rows = self.db.execute_query(
            "SELECT inventory_number FROM instances WHERE inventory_number IS NOT NULL",
            fetch_all=True
        ) or []
        for row in rows:
            v = row['inventory_number']
            if v:
                self._inv_cache.add(str(v).strip())

        rows = self.db.execute_query(
            "SELECT serial_number FROM instances WHERE serial_number IS NOT NULL",
            fetch_all=True
        ) or []
        for row in rows:
            v = row['serial_number']
            if v:
                self._serial_cache.add(str(v).strip())

        rows = self.db.execute_query(
            "SELECT sku FROM nomenclatures WHERE sku IS NOT NULL",
            fetch_all=True
        ) or []
        self._existing_skus = {row['sku'] for row in rows}

        rows = self.db.execute_query(
            "SELECT id, name FROM nomenclatures WHERE is_deleted = 0 OR is_deleted IS NULL",
            fetch_all=True
        ) or []
        for row in rows:
            if row['name']:
                self._nom_cache[row['name'].strip().lower()] = row['id']

        logger.info(
            f"Cache loaded: {len(self._inv_cache)} inv_numbers, "
            f"{len(self._serial_cache)} serials, "
            f"{len(self._nom_cache)} nomenclatures"
        )

    def _get_or_create_nomenclature(
        self,
        name: str,
        model: str = None,
        description: str = None,
        unit: str = 'шт',
        accounting_type: str = 'individual',
        category_code: str = None,
        category_id_override: int = None
    ) -> int:
        """Get existing or create new nomenclature. Returns id."""
        if not name or not name.strip():
            return None

        key = name.strip().lower()
        if key in self._nom_cache:
            self.stats['nomenclatures_skipped'] += 1
            return self._nom_cache[key]

        # Determine category_id: explicit override > code lookup > default
        if category_id_override:
            category_id = category_id_override
        else:
            category_id = getattr(self, '_default_category_id', 1)
            if category_code:
                row = self.db.execute_query(
                    "SELECT id FROM categories WHERE code = ? LIMIT 1",
                    (str(category_code).strip(),),
                    fetch_all=False
                )
                if row:
                    category_id = row['id']

        # Generate SKU using proper Cyrillic-prefix format
        sku = generate_sku_from_name(name.strip(), self._existing_skus)

        # Create nomenclature
        has_serial = 1 if accounting_type == 'individual' else 0
        self.db.execute_query(
            """
            INSERT INTO nomenclatures
                (sku, name, description, model, category_id, accounting_type, unit,
                 has_serial_numbers, is_active, is_deleted, created_at, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1, 0, CURRENT_TIMESTAMP, ?)
            """,
            (sku, name.strip(), description, model, category_id,
             accounting_type, unit, has_serial, self.user_id)
        )
        nom_id = self.db.cursor.lastrowid

        self._nom_cache[key] = nom_id
        self.stats['nomenclatures_created'] += 1
        return nom_id

    def _is_duplicate(self, inventory_number: str, serial_number: str) -> bool:
        """Check if instance already exists."""
        if inventory_number and str(inventory_number).strip() in self._inv_cache:
            return True
        if serial_number and str(serial_number).strip() in self._serial_cache:
            return True
        return False

    def _create_instance(
        self,
        nomenclature_id: int,
        inventory_number: str,
        old_inventory_number: str,
        serial_number: str,
        status: str,
        location_name: str,
        employee_name: str,
        purchase_date,
        purchase_price,
        accounting_type: str = 'individual'
    ):
        """Create instance record."""
        if not nomenclature_id:
            return False

        inv_num = str(inventory_number).strip() if inventory_number else None
        old_inv_num = str(old_inventory_number).strip() if old_inventory_number else None
        ser_num = str(serial_number).strip() if serial_number else None

        if self._is_duplicate(inv_num, ser_num):
            self.stats['instances_skipped'] += 1
            return False

        # Generate inventory number if missing (required for view URL)
        if not inv_num:
            row = self.db.execute_query("SELECT COUNT(*) as cnt FROM instances", fetch_all=False)
            cnt = row['cnt'] if row else 0
            inv_num = f"IMP-{cnt + 1:06d}"
            while inv_num in self._inv_cache:
                cnt += 1
                inv_num = f"IMP-{cnt:06d}"

        # Map status
        mapped_status = STATUS_MAP.get(
            str(status).strip() if status else '',
            'in_stock'
        )

        # Find or create location_id
        location_id = None
        if location_name and str(location_name).strip():
            loc = str(location_name).strip()
            row = self.db.execute_query(
                "SELECT id FROM locations WHERE name = ? LIMIT 1", (loc,), fetch_all=False
            )
            if row:
                location_id = row['id']
            else:
                loc_code = re.sub(r'[^A-Za-zА-Яа-я0-9]', '', loc).upper()[:20] or 'LOC'
                cnt_row = self.db.execute_query(
                    "SELECT COUNT(*) as cnt FROM locations WHERE code LIKE ?",
                    (loc_code + '%',), fetch_all=False
                )
                cnt = cnt_row['cnt'] if cnt_row else 0
                if cnt:
                    loc_code = f"{loc_code[:16]}{cnt:04d}"
                self.db.execute_query(
                    "INSERT INTO locations (code, name, type, is_active, created_at) VALUES (?, ?, 'office', 1, CURRENT_TIMESTAMP)",
                    (loc_code, loc)
                )
                location_id = self.db.cursor.lastrowid

        # Find employee_id
        employee_id = None
        if employee_name and str(employee_name).strip():
            emp = str(employee_name).strip()
            parts = emp.split()
            if parts:
                row = self.db.execute_query(
                    "SELECT id FROM employees WHERE last_name = ? LIMIT 1",
                    (parts[0],), fetch_all=False
                )
                if row:
                    employee_id = row['id']

        # Parse purchase_date
        pd_str = None
        if purchase_date:
            if isinstance(purchase_date, datetime):
                pd_str = purchase_date.strftime('%Y-%m-%d')
            else:
                try:
                    pd_str = str(purchase_date)[:10]
                except Exception:
                    pd_str = None

        # Parse price
        price = None
        if purchase_price:
            try:
                price = float(
                    str(purchase_price).replace(',', '.').replace(' ', '')
                )
            except (ValueError, TypeError):
                price = None

        # Create instance
        self.db.execute_query(
            """
            INSERT INTO instances
                (nomenclature_id, inventory_number, old_inventory_number,
                 serial_number, status, condition, location_id, employee_id,
                 purchase_date, purchase_price, created_at, created_by)
            VALUES (?, ?, ?, ?, ?, 'good', ?, ?, ?, ?, CURRENT_TIMESTAMP, ?)
            """,
            (nomenclature_id, inv_num, old_inv_num, ser_num,
             mapped_status, location_id, employee_id, pd_str, price, self.user_id)
        )

        # Update caches
        if inv_num:
            self._inv_cache.add(inv_num)
        if ser_num:
            self._serial_cache.add(ser_num)

        self.stats['instances_created'] += 1
        return True

    def _import_file(self, path: str, fmt: str):
        """Import a single Excel file."""
        logger.info(f"Importing {os.path.basename(path)} (format={fmt})")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

        # For 'instances' format use the named sheet if available
        if fmt == 'instances' and 'Шаблон' in wb.sheetnames:
            ws = wb['Шаблон']
        else:
            ws = wb.active

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        logger.info(f"  Rows to process: {len(rows)}")

        for i, row in enumerate(rows):
            try:
                if fmt == 'standard':
                    self._process_standard_row(row)
                elif fmt == 'tool':
                    self._process_tool_row(row)
                elif fmt == 'instances':
                    self._process_instances_row(row)
            except Exception as e:
                logger.error(f"  Row {i+2} error: {e}")
                self.stats['errors'].append(
                    f"{os.path.basename(path)} row {i+2}: {str(e)}"
                )

        wb.close()

    def _process_standard_row(self, row):
        """Process row from inventory/equipment/assets/consumable files.

        Columns:
            0  id
            1  name
            2  subcategory
            3  asset_type
            4  year
            5  old_inventory_number
            6  inventory_number
            7  serial_number
            8  model
            9  location
            10 department
            11 responsible
            12 description
            13 status
            14 value
            ...
            30 category_code  (approximately)
        """
        if not row or not row[1]:  # name is required
            return

        name = str(row[1]).strip()
        asset_type = row[3]
        old_inv = row[5]
        inv_num = row[6]
        serial = row[7]
        model = row[8]
        location = row[9]
        responsible = row[11]
        description = row[12]
        status = row[13]
        value = row[14]
        category_code = row[30] if len(row) > 30 else None

        # Determine accounting_type
        acc_type = (
            'individual'
            if asset_type in ('inventory', 'equipment')
            else 'quantitative'
        )

        nom_id = self._get_or_create_nomenclature(
            name=name,
            model=str(model) if model else None,
            description=str(description) if description else None,
            unit='шт',
            accounting_type=acc_type,
            category_code=str(category_code) if category_code else None
        )

        self._create_instance(
            nomenclature_id=nom_id,
            inventory_number=inv_num,
            old_inventory_number=old_inv,
            serial_number=serial,
            status=status,
            location_name=str(location) if location else None,
            employee_name=str(responsible) if responsible else None,
            purchase_date=None,
            purchase_price=value
        )

    def _process_tool_row(self, row):
        """Process row from tool.xlsx.

        Columns:
            0  Бр/шт тип подш  (skip)
            1  Инв. Номер       -> inventory_number
            2  Наименование     -> name
            3  Зав/опер номер   -> serial_number
            4  Функция          -> description
            5  Паронамендающие  (skip)
            6  Инвентар         -> old_inventory_number
            7  Дата выдачи      -> purchase_date
            8  Объем            (skip)
        """
        if not row or not row[2]:  # name required
            return

        name = str(row[2]).strip()
        inv_num = row[1]
        serial = row[3]
        description = row[4]
        old_inv = row[6]
        purchase_date = row[7]

        nom_id = self._get_or_create_nomenclature(
            name=name,
            description=str(description) if description else None,
            unit='шт',
            accounting_type='individual'
        )

        self._create_instance(
            nomenclature_id=nom_id,
            inventory_number=inv_num,
            old_inventory_number=old_inv,
            serial_number=serial,
            status='in_stock',
            location_name=None,
            employee_name=None,
            purchase_date=purchase_date,
            purchase_price=None
        )

    # ------------------------------------------------------------------ helpers

    def _get_category_id_by_code(self, code: str, fallback: int = 1) -> int:
        row = self.db.execute_query(
            "SELECT id FROM categories WHERE code = ? LIMIT 1", (code,), fetch_all=False
        )
        return row['id'] if row else fallback

    def _get_default_warehouse_id(self) -> int:
        row = self.db.execute_query(
            "SELECT id FROM warehouses WHERE is_active = 1 ORDER BY id LIMIT 1",
            fetch_all=False
        )
        return row['id'] if row else 1

    def _load_stock_cache(self) -> set:
        """Cache existing (nomenclature_id, warehouse_id) pairs in stocks."""
        rows = self.db.execute_query(
            "SELECT nomenclature_id, warehouse_id FROM stocks", fetch_all=True
        ) or []
        return {(r['nomenclature_id'], r['warehouse_id']) for r in rows}

    @staticmethod
    def _extract_serial(raw_name: str):
        """Extract serial number embedded in asset name.

        Patterns recognised:
          "... сн043172220348"      → serial="043172220348"
          "... .сн-HFE5305320VA"   → serial="HFE5305320VA"
          "... сн-SY300-330"       → serial="SY300-330"
        Returns (clean_name, serial_or_None).
        """
        pattern = r'[.\s]сн[-.\s]?([A-Za-z0-9][A-Za-z0-9.\-/]+)'
        m = re.search(pattern, raw_name, re.IGNORECASE)
        if m:
            serial = m.group(1).strip('.').strip()
            clean = raw_name[:m.start()].strip(' .')
            return clean, serial
        return raw_name.strip(), None

    @staticmethod
    def _is_account_code_row(cell_value) -> bool:
        """True for subtotal rows like '01', '01.01', '10.01.1' etc."""
        if not cell_value:
            return True
        s = str(cell_value).strip()
        if not s:
            return True
        # pure account codes are short digit+dot strings
        return bool(re.match(r'^\d[\d.]*$', s)) and len(s) <= 10

    def _create_stock(self, nomenclature_id: int, quantity: float, unit_price: float):
        """Create or update stock record for a quantitative item."""
        wid = getattr(self, '_default_warehouse_id', 1)
        key = (nomenclature_id, wid)
        if key in self._stock_cache:
            self.stats['instances_skipped'] += 1
            return
        qty_int = int(round(quantity)) if quantity else 0
        self.db.execute_query(
            """
            INSERT OR IGNORE INTO stocks
                (nomenclature_id, warehouse_id, quantity, reserved_quantity,
                 created_at, updated_at)
            VALUES (?, ?, ?, 0, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            """,
            (nomenclature_id, wid, qty_int)
        )
        self._stock_cache.add(key)
        self.stats['instances_created'] += 1

    # ---------------------------------------------------------- ОСВ importers

    def _import_osv_file(self, path: str, fmt: str):
        """Import an ОСВ (trial-balance) file.
        Files may be true .xlsx or legacy .xls saved with .xlsx extension.
        Tries openpyxl first, falls back to xlrd for legacy format.
        """
        logger.info(f"Importing OSV {os.path.basename(path)} (format={fmt})")
        rows = self._read_spreadsheet_rows(path)
        logger.info(f"  Total rows: {len(rows)}")

        if fmt == 'osv01':
            self._process_osv01_rows(rows, os.path.basename(path))
        else:
            self._process_osv10_rows(rows, os.path.basename(path))

    @staticmethod
    def _read_spreadsheet_rows(path: str) -> list:
        """Read all rows from xlsx or legacy xls file. Returns list of value tuples."""
        # Try openpyxl first (true xlsx)
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
            wb.close()
            return rows
        except Exception:
            pass
        # Fall back to xlrd (legacy .xls or mislabelled files)
        import xlrd
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_index(0)
        rows = []
        for ri in range(ws.nrows):
            rows.append(tuple(ws.cell_value(ri, ci) for ci in range(ws.ncols)))
        return rows

    def _process_osv01_rows(self, rows, filename: str):
        """счёт 01 — Основные средства.

        Layout (7 cols):
          col 0  name (+ embedded serial after 'сн')
          col 1  debet_start
          col 2  credit_start
          col 3  debet_turnover
          col 4  credit_turnover
          col 5  debet_end   ← balance / value at end of period
          col 6  credit_end
        Data starts when col 0 is a long name (not an account code).
        Rows with debet_end = 0 → written off, still import as written_off.
        """
        category_id = getattr(self, '_category_id_01', 1)

        for ri, row in enumerate(rows):
            try:
                if not row or len(row) < 6:
                    continue
                raw_name = row[0]
                if self._is_account_code_row(raw_name):
                    continue

                name, serial = self._extract_serial(str(raw_name).strip())
                if not name:
                    continue

                debet_end = row[5]
                try:
                    value = float(debet_end) if debet_end else 0.0
                except (TypeError, ValueError):
                    value = 0.0

                # If serial already in cache → duplicate
                if serial and str(serial).strip() in self._serial_cache:
                    self.stats['instances_skipped'] += 1
                    continue

                nom_id = self._get_or_create_nomenclature(
                    name=name,
                    unit='шт',
                    accounting_type='individual',
                    category_id_override=category_id
                )
                if not nom_id:
                    continue

                status = 'in_stock' if value > 0 else 'written_off'
                self._create_instance(
                    nomenclature_id=nom_id,
                    inventory_number=None,
                    old_inventory_number=None,
                    serial_number=serial,
                    status=status,
                    location_name=None,
                    employee_name=None,
                    purchase_date=None,
                    purchase_price=value if value > 0 else None
                )
            except Exception as e:
                logger.error(f"  {filename} row {ri}: {e}")
                self.stats['errors'].append(f"{filename} row {ri}: {str(e)}")

    def _process_osv10_rows(self, rows, filename: str):
        """счёт 10 — Материалы.

        Layout (8 cols), data comes in PAIRS:
          БУ  row: col 0=name, col 1='БУ',  col 6=debet_end (value)
          Кол row: col 0='',   col 1='Кол.', col 6=qty_end

        Skip items where qty_end = 0 and debet_end = 0 (fully consumed).
        Still create nomenclature even if qty=0 (might be re-ordered later).
        """
        category_id = getattr(self, '_category_id_10', 2)
        i = 0
        while i < len(rows):
            row = rows[i]
            # Find БУ row: col1 == 'БУ' and col0 is a non-empty name
            if (row and len(row) >= 7
                    and str(row[1] or '').strip() == 'БУ'
                    and not self._is_account_code_row(row[0])):
                try:
                    name = str(row[0]).strip()
                    debet_end = row[6]
                    try:
                        value = float(debet_end) if debet_end else 0.0
                    except (TypeError, ValueError):
                        value = 0.0

                    # Next row should be Кол.
                    qty = 0.0
                    if i + 1 < len(rows):
                        kol_row = rows[i + 1]
                        if (kol_row and len(kol_row) >= 7
                                and str(kol_row[1] or '').strip() == 'Кол.'):
                            try:
                                qty = float(kol_row[6]) if kol_row[6] else 0.0
                            except (TypeError, ValueError):
                                qty = 0.0
                            i += 1  # skip the Кол. row

                    # PPE items should be individual even in счёт 10
                    _ppe_kws = [
                        'сапог', 'берц', 'перчат', 'каска', 'жилет', 'костюм',
                        'защит', 'маска', 'наушник', 'ботин', 'кроссовк',
                    ]
                    _name_lower = name.lower()
                    _acc_type = (
                        'individual'
                        if any(kw in _name_lower for kw in _ppe_kws)
                        else 'quantitative'
                    )
                    nom_id = self._get_or_create_nomenclature(
                        name=name,
                        unit='шт',
                        accounting_type=_acc_type,
                        category_id_override=category_id
                    )
                    if nom_id and qty > 0:
                        unit_price = value / qty if qty > 0 else 0.0
                        self._create_stock(nom_id, qty, unit_price)
                    elif nom_id and qty == 0:
                        # No stock but nomenclature created — count as skipped instance
                        self.stats['instances_skipped'] += 1

                except Exception as e:
                    logger.error(f"  {filename} row {i}: {e}")
                    self.stats['errors'].append(f"{filename} row {i}: {str(e)}")
            i += 1

    # -------------------------------------------------------------- existing

    def _process_instances_row(self, row):
        """Process row from Экземпляры.xlsx (sheet Шаблон).

        Columns:
            0  Старый инв. номер   -> old_inventory_number
            1  Инв. номер          -> inventory_number  (DEDUP KEY)
            2  Номенклатура/Комплект -> name
            3  Комплектующие       (skip)
            4  Серийный номер      -> serial_number  (DEDUP KEY)
            5  Статус              -> status
            6  Местоположение      -> location
            7  Сотрудник           -> employee
            8  Дата покупки        -> purchase_date
            9  Цена                -> purchase_price
            10 Количество          (informational)
        """
        if not row or not row[2]:  # name required
            return

        old_inv = row[0]
        inv_num = row[1]
        name = str(row[2]).strip()
        serial = row[4]
        status = row[5]
        location = row[6]
        employee = row[7]
        purchase_date = row[8]
        price = row[9]

        nom_id = self._get_or_create_nomenclature(
            name=name,
            unit='шт',
            accounting_type='individual'
        )

        self._create_instance(
            nomenclature_id=nom_id,
            inventory_number=inv_num,
            old_inventory_number=old_inv,
            serial_number=serial,
            status=status,
            location_name=str(location) if location else None,
            employee_name=str(employee) if employee else None,
            purchase_date=purchase_date,
            purchase_price=price
        )
